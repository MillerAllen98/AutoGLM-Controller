import subprocess
import os
import threading
import tkinter as tk
from tkinter import font, ttk, messagebox, filedialog
from datetime import datetime
import time
import csv
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from phone_agent import PhoneAgent
from phone_agent.agent import AgentConfig
from phone_agent.agent_ios import IOSAgentConfig, IOSPhoneAgent
from phone_agent.model import ModelConfig
from phone_agent.device_factory import DeviceType, set_device_type

try:
    from tkmacosx import Button as MacButton
except ImportError:
    MacButton = tk.Button

PLATFORMS = {
    "ZhipuAI-AutoGLM": {"base_url": "https://open.bigmodel.cn/api/paas/v4", "model": "autoglm-phone"},
    "ModelScope-Agent": {"base_url": "https://api-inference.modelscope.cn/v1", "model": "ZhipuAI/AutoGLM-Phone-9B"}
}


class TextRedirector:
    def __init__(self, widget, parent_instance):
        self.widget = widget
        self.parent = parent_instance
        self.buffer = ""

    def write(self, string):
        self.buffer += string
        self.widget.insert(tk.END, string)
        self.widget.see(tk.END)

    def flush(self):
        pass

    def get_and_clear_buffer(self):
        text = self.buffer
        self.buffer = ""
        return text


class WindowsClassicAutoGLM:
    def __init__(self, root):
        self.root = root
        self.root.title("AutoGLM UI Automation Controller")
        self.root.geometry("1200x950")

        self.iproxy_process = None
        self.agent = None
        self.current_thread = None
        self.is_manual_stop = False
        self.progress_val = tk.DoubleVar()
        self.excel_data = None

        self.stop_timer = None
        self.FORCE_STOP_TIMEOUT = 60

        self.current_case_step_count = 0
        self.max_steps_for_current_case = 0
        self._current_cid = "UNKNOWN"

        self.clr_bg = "#d4d0c8"
        self.font_main = ("Tahoma", 12)
        self.font_bold = ("Tahoma", 12, "bold")
        self.font_header = ("Tahoma", 14, "bold")
        self.font_console = ("Courier New", 12)
        self.root.configure(bg=self.clr_bg)

        self._setup_styles()
        self._build_interface()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use('classic')
        style.configure("TLabel", background=self.clr_bg, font=self.font_main)
        style.configure("TNotebook", background=self.clr_bg)
        style.configure("TNotebook.Tab", background=self.clr_bg, font=self.font_main, padding=[10, 5])

    def _build_interface(self):
        menu_bar = tk.Frame(self.root, bg=self.clr_bg, bd=1, relief="raised")
        menu_bar.pack(fill=tk.X)
        for text in ["文件(F)", "执行(R)", "设置(S)", "帮助(H)"]:
            tk.Label(menu_bar, text=text, bg=self.clr_bg, padx=12, pady=5, font=self.font_main).pack(side=tk.LEFT)

        main_container = tk.Frame(self.root, bg=self.clr_bg, padx=15, pady=15)
        main_container.pack(fill=tk.BOTH, expand=True)

        left_panel = tk.LabelFrame(main_container, text=" 环境配置 ", bg=self.clr_bg, font=self.font_header, padx=15, pady=15)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 15))
        left_panel.config(width=320)
        left_panel.pack_propagate(False)

        def add_row(parent, text):
            f = tk.Frame(parent, bg=self.clr_bg, pady=8)
            f.pack(fill=tk.X)
            tk.Label(f, text=text, bg=self.clr_bg, width=6, anchor="w").pack(side=tk.LEFT)
            return f

        self.os_combo = ttk.Combobox(add_row(left_panel, "平台:"), values=["Android", "iOS"], state="readonly", width=15)
        self.os_combo.current(1)
        self.os_combo.pack(side=tk.LEFT)

        self.plat_combo = ttk.Combobox(add_row(left_panel, "引擎:"), values=list(PLATFORMS.keys()), state="readonly", width=15)
        self.plat_combo.current(0)
        self.plat_combo.pack(side=tk.LEFT)

        self.key_entry = tk.Entry(add_row(left_panel, "密钥:"), bg="white", relief="sunken", bd=2)
        self.key_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.strict_mode_var = tk.BooleanVar(value=True)
        tk.Checkbutton(left_panel, text="严格模式（禁止AI自由发挥、只按步骤执行）", variable=self.strict_mode_var,
                       bg=self.clr_bg, font=("Tahoma", 10), selectcolor=self.clr_bg, wraplength=250, justify="left").pack(pady=20)

        btn_cfg = {"width": 20, "height": 2, "relief": "raised", "bd": 4, "font": self.font_bold}
        self.btn_go = tk.Button(left_panel, text="开始执行 (RUN)", command=self.handle_start, **btn_cfg)
        self.btn_go.pack(pady=10)
        self.btn_stop = tk.Button(left_panel, text="停止 (STOP)", state="disabled", command=self.handle_stop, **btn_cfg)
        self.btn_stop.pack(pady=10)

        self.btn_export = tk.Button(left_panel, text="导出报告 (EXPORT)", state="disabled", command=self.export_report, **btn_cfg)
        self.btn_export.pack(pady=10)

        right_panel = tk.Frame(main_container, bg=self.clr_bg)
        right_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        case_container = tk.LabelFrame(right_panel, text=" ui自动化测试用例 ", bg=self.clr_bg, font=self.font_header, padx=10, pady=10)
        case_container.pack(fill=tk.X, pady=(0, 10))
        self.tab_control = ttk.Notebook(case_container)
        self.tab_control.pack(fill=tk.X)
        self.tab_control.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        self.tab_text = tk.Frame(self.tab_control, bg=self.clr_bg, pady=5)
        self.prompt_input = tk.Text(self.tab_text, height=5, font=self.font_console, bd=2, relief="sunken")
        self.prompt_input.pack(fill=tk.X)
        self.tab_control.add(self.tab_text, text=" 文本框输入 ")

        self.tab_excel = tk.Frame(self.tab_control, bg=self.clr_bg, pady=10)
        btn_fm = tk.Frame(self.tab_excel, bg=self.clr_bg)
        btn_fm.pack(fill=tk.X)
        tk.Button(btn_fm, text="下载用例模板", command=self.download_template_csv).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_fm, text="上传测试用例", command=self.upload_excel_logic).pack(side=tk.LEFT, padx=5)
        self.file_label = tk.Label(self.tab_excel, text="未选择文件", bg=self.clr_bg, fg="blue")
        self.file_label.pack(side=tk.LEFT, padx=10)
        self.tab_control.add(self.tab_excel, text=" Excel 模板上传 ")

        monitor_container = tk.Frame(right_panel, bg=self.clr_bg)
        monitor_container.pack(fill=tk.BOTH, expand=True)

        case_frame = tk.LabelFrame(monitor_container, text=" 用例执行状态 ", bg=self.clr_bg, font=self.font_header)
        case_frame.place(relx=0, rely=0, relwidth=0.49, relheight=1)
        self.case_area = tk.Text(case_frame, bg="black", fg="#00FFFF", font=self.font_console, bd=2, relief="sunken")
        self.case_area.pack(fill=tk.BOTH, expand=True)

        log_frame = tk.LabelFrame(monitor_container, text=" 实时监控日志 (Detail) ", bg=self.clr_bg, font=self.font_header)
        log_frame.place(relx=0.51, rely=0, relwidth=0.49, relheight=1)
        self.log_area = tk.Text(log_frame, bg="black", fg="#00ff00", font=self.font_console, bd=2, relief="sunken")
        self.log_area.pack(fill=tk.BOTH, expand=True)

        self.status_bar = tk.Frame(self.root, bg=self.clr_bg, bd=1, relief="sunken", height=45)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        self.progress_bar = ttk.Progressbar(self.status_bar, orient="horizontal", length=300, variable=self.progress_val)
        self.progress_bar.pack(side=tk.RIGHT, padx=15, pady=8)
        self.status_text_lbl = tk.Label(self.status_bar, text="未启动", bg=self.clr_bg, font=self.font_bold)
        self.status_text_lbl.pack(side=tk.LEFT, padx=10)

    def _on_tab_changed(self, event):
        active_tab = self.tab_control.index(self.tab_control.select())
        if active_tab == 0:
            self.excel_data = None
            self.file_label.config(text="未选择文件")
        else:
            self.prompt_input.delete("1.0", tk.END)

    def download_template_csv(self):
        path = filedialog.asksaveasfilename(title="保存用例模板", initialfile="UI自动化用例模板.csv", defaultextension=".csv")
        if path:
            try:
                headers = ["用例编号", "用例标题", "前置条件", "测试步骤", "预期结果", "优先级", "备注"]
                with open(path, 'w', newline='', encoding='utf_8_sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerow(["CASE_001", "朋友圈点赞", "已登录微信", "1.点击发现\n2.点击朋友圈",
                                     "1.进入发现页\n2.成功进入朋友圈", "P0", "无"])
                messagebox.showinfo("成功", "模板下载成功")
            except Exception as e:
                messagebox.showerror("错误", str(e))

    def upload_excel_logic(self):
        try:
            import pandas as pd
            path = filedialog.askopenfilename(filetypes=[("Excel/CSV", "*.xlsx *.csv")])
            if path:
                df = pd.read_excel(path) if path.endswith('.xlsx') else pd.read_csv(path)
                self.excel_data = df.to_dict('records')
                self.file_label.config(text=f"已加载: {len(self.excel_data)} 条用例")
        except Exception as e:
            messagebox.showerror("错误", "上传需安装 pandas\n" + str(e))

    def _ensure_iproxy(self):
        if self.os_combo.get() == "iOS":
            try:
                if self.iproxy_process is None or self.iproxy_process.poll() is not None:
                    self.iproxy_process = subprocess.Popen(["iproxy", "8100", "8100"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    time.sleep(2.5)
                return True
            except Exception as e:
                print(f"iproxy 启动失败: {e}")
                return False
        return True

    def _initialize_agent_instance(self, api_key):
        if self.agent is not None:
            return
        try:
            plat = PLATFORMS[self.plat_combo.get()]
            device_type_str = "ios" if self.os_combo.get() == "iOS" else "adb"
            dt = DeviceType.IOS if device_type_str == "ios" else DeviceType.ADB

            model_config = ModelConfig(
                base_url=plat["base_url"],
                model_name=plat["model"],
                api_key=api_key,
                lang="cn"
            )

            strict_system_prompt = (
                "你是一个完全没有智能修复能力的机械执行机器人。\n"
                "核心铁律（绝对不可违反）：\n"
                "1. 严禁任何滑动、滚动、swipe、scroll、上滑、下拉等屏幕移动操作。\n"
                "2. 只能点击文本100%完全精确匹配的元素。\n"
                "3. 元素必须在当前屏幕直接可见，不可见 = 立即输出 FAILED。\n"
                "4. 遇到弹窗、权限、广告、登录页，一律不点击，直接输出 FAILED。\n"
                "5. 每步最多 2 次原子动作，超出必须立即停止并输出 FAILED。\n"
                "6. 每步执行结束后，请在最后一行明确写：'本步执行成功' 或 '本步执行失败'。"
            )

            if self.strict_mode_var.get():
                model_config.system_prompt = strict_system_prompt

            if dt == DeviceType.IOS:
                agent_config = IOSAgentConfig(
                    max_steps=100,
                    wda_url="http://localhost:8100",
                    device_id=None,
                    verbose=True,
                    lang="cn"
                )
                self.agent = IOSPhoneAgent(model_config=model_config, agent_config=agent_config)
            else:
                agent_config = AgentConfig(
                    max_steps=100,
                    device_id=None,
                    verbose=True,
                    lang="cn"
                )
                self.agent = PhoneAgent(model_config=model_config, agent_config=agent_config)

            print(f"✅ Agent 初始化成功（严格模式: {'启用' if self.strict_mode_var.get() else '关闭'}）")
        except Exception as e:
            print(f"❌ Agent 初始化失败: {e}")
            raise

    def _is_step_success_by_text(self, text):
        if not text:
            return False
        text_lower = text.lower()
        success_keywords = ["成功", "进入", "打开", "点击成功", "完成", "已进入", "已打开", "已点击", "本步执行成功"]
        fail_keywords = ["失败", "未找到", "没有", "无法", "找不到", "不存在", "未发现", "本步执行失败", "failed"]

        if any(kw in text_lower for kw in fail_keywords):
            return False
        if any(kw in text_lower for kw in success_keywords):
            return True
        return False

    def _reset_ui_to_stopped(self):
        self.btn_go.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.btn_export.config(state="normal")
        self.status_text_lbl.config(text="已强制停止", fg="red")

    def _force_stop_action(self):
        print(f"⚠️ 等待超时 ({self.FORCE_STOP_TIMEOUT}秒)，执行强制停止...")
        self.is_manual_stop = True

        if self.os_combo.get() == "iOS":
            try:
                subprocess.run(["pkill", "-f", "iproxy.*8100"], timeout=3)
                print("🔪 [iOS] 已强制终止所有 iproxy 8100 进程")
            except:
                pass
            self.iproxy_process = None
        elif self.os_combo.get() == "Android":
            try:
                subprocess.run(["adb", "kill-server"], timeout=5)
                print("🔪 [Android] 已终止 ADB Server")
            except:
                pass

        self.agent = None
        print("🧹 Agent 实例已强制销毁")
        self.root.after(0, self._reset_ui_to_stopped)

    def handle_stop(self):
        self.is_manual_stop = True
        self.status_text_lbl.config(text="正在停止 (等待当前用例完成)...", fg="orange")
        if self.stop_timer:
            self.stop_timer.cancel()
        self.stop_timer = threading.Timer(self.FORCE_STOP_TIMEOUT, self._force_stop_action)
        self.stop_timer.start()

    def handle_start(self):
        active_tab = self.tab_control.index(self.tab_control.select())
        cases = []
        if active_tab == 0:
            txt = self.prompt_input.get("1.0", tk.END).strip()
            if not txt:
                return
            cases = [{"用例编号": "TEXT_MODE", "测试步骤": txt}]
        else:
            if not self.excel_data:
                return
            cases = self.excel_data

        api_key = self.key_entry.get().strip()
        if not api_key:
            messagebox.showerror("错误", "缺少 Key")
            return

        self.is_manual_stop = False
        self.progress_val.set(0)
        self.status_text_lbl.config(text="正在进行...", fg="blue")
        self.case_area.delete("1.0", tk.END)
        self.log_area.delete("1.0", tk.END)
        self.btn_go.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.btn_export.config(state="disabled")

        if self.stop_timer:
            self.stop_timer.cancel()

        self.current_thread = threading.Thread(target=self._run_task, args=(cases, api_key), daemon=True)
        self.current_thread.start()

    def export_report(self):
        result_text = self.case_area.get("1.0", tk.END).strip()
        if not result_text:
            messagebox.showwarning("警告", "没有执行结果可导出")
            return

        save_path = filedialog.asksaveasfilename(
            title="保存测试报告",
            defaultextension=".xlsx",
            initialfile=f"AutoGLM_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if not save_path:
            return

        try:
            lines = result_text.split('\n')

            total_cases = 0
            passed_cases = 0
            case_results = []

            for line in lines:
                if "用例执行成功" in line or "用例失败" in line:
                    total_cases += 1
                    if "成功" in line:
                        passed_cases += 1
                    parts = line.split()
                    cid = next((p for p in parts if p.startswith("CASE_")), "UNKNOWN")
                    status = "通过" if "成功" in line else "失败"
                    case_results.append({"用例编号": cid, "执行结果": status})

            failed_cases = total_cases - passed_cases
            pass_rate = (passed_cases / total_cases * 100) if total_cases > 0 else 0

            wb = Workbook()
            ws_summary = wb.active
            ws_summary.title = "测试报告"

            # 样式定义
            title_font = Font(size=20, bold=True, color="FFFFFF")
            header_font = Font(bold=True, color="FFFFFF")
            data_font = Font(size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            fail_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # 标题
            ws_summary['A1'] = "AutoGLM UI 自动化测试报告"
            ws_summary['A1'].font = title_font
            ws_summary['A1'].alignment = Alignment(horizontal="center")
            ws_summary.merge_cells('A1:F1')
            ws_summary.row_dimensions[1].height = 40

            # 概览数据
            data = [
                ["总用例数", total_cases],
                ["通过用例", passed_cases],
                ["失败用例", failed_cases],
                ["通过率", f"{pass_rate:.2f}%"],
                ["执行时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["平台", self.os_combo.get()],
                ["引擎", self.plat_combo.get()]
            ]

            for i, (k, v) in enumerate(data, start=3):
                ws_summary[f'A{i}'] = k
                ws_summary[f'B{i}'] = v
                ws_summary[f'A{i}'].font = data_font
                ws_summary[f'B{i}'].font = data_font
                ws_summary[f'A{i}'].alignment = Alignment(horizontal="right")
                if k == "通过用例":
                    ws_summary[f'B{i}'].fill = pass_fill
                    ws_summary[f'B{i}'].font = Font(color="FFFFFF", bold=True)
                elif k == "失败用例":
                    ws_summary[f'B{i}'].fill = fail_fill
                    ws_summary[f'B{i}'].font = Font(color="FFFFFF", bold=True)

            # 【关键修复】安全自动调整列宽（跳过 MergedCell）
            column_widths = {}
            for row in ws_summary.iter_rows():
                for cell in row:
                    if cell.value is not None and not getattr(cell, 'merged', False):
                        col_letter = cell.column_letter
                        cell_length = len(str(cell.value))
                        column_widths[col_letter] = max(column_widths.get(col_letter, 8), cell_length + 2)

            for col, width in column_widths.items():
                ws_summary.column_dimensions[col].width = min(width, 50)

            # 用例明细表
            ws_detail = wb.create_sheet("用例明细")
            df_detail = pd.DataFrame(case_results)
            for r in dataframe_to_rows(df_detail, index=False, header=True):
                ws_detail.append(r)

            # 美化明细表
            for cell in ws_detail[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row in ws_detail.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
                    if cell.column_letter == 'B':  # 执行结果列
                        if cell.value == "通过":
                            cell.fill = pass_fill
                            cell.font = Font(color="FFFFFF")
                        else:
                            cell.fill = fail_fill
                            cell.font = Font(color="FFFFFF")

            # 明细表列宽
            ws_detail.column_dimensions['A'].width = 25
            ws_detail.column_dimensions['B'].width = 15

            wb.save(save_path)
            messagebox.showinfo("成功", f"专业测试报告已导出：\n{save_path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _run_task(self, cases, api_key):
        if not self._ensure_iproxy():
            self.root.after(0, lambda: self.status_text_lbl.config(text="iproxy 启动失败", fg="red"))
            self.root.after(0, self._reset_ui_to_stopped)
            return

        try:
            self._initialize_agent_instance(api_key)
        except Exception as e:
            self.root.after(0, lambda: self.status_text_lbl.config(text="Agent 初始化失败", fg="red"))
            self.root.after(0, self._reset_ui_to_stopped)
            return

        original_stdout = sys.stdout
        redirector = TextRedirector(self.log_area, self)
        sys.stdout = redirector

        try:
            total = len(cases)
            for idx, row in enumerate(cases):
                if self.is_manual_stop or self.agent is None:
                    print("⚠️ 检测到停止信号，退出循环")
                    break

                progress_percent = (idx / total) * 100
                self.root.after(0, lambda p=progress_percent, c=idx + 1, t=total: [
                    self.progress_val.set(p),
                    self.status_text_lbl.config(text=f"正在进行: 用例 {c}/{t}", fg="blue")
                ])

                cid = str(row.get("用例编号", f"CASE_{idx + 1}"))
                self._current_cid = cid
                steps_raw = str(row.get("测试步骤", ""))
                steps_list = [s.strip() for s in steps_raw.split('\n') if s.strip()]

                self.current_case_step_count = 0
                self.max_steps_for_current_case = len(steps_list)

                print(f"\n{'=' * 40}")
                print(f"开始执行用例: {cid} （共 {len(steps_list)} 步）")
                print(f"{'=' * 40}")

                redirector.buffer = ""

                for step_idx, single_step in enumerate(steps_list):
                    if self.is_manual_stop or self.agent is None:
                        break

                    per_step_prompt = (
                        f"【当前第 {step_idx + 1}/{len(steps_list)} 步】\n"
                        f"指令：{single_step}\n\n"
                        "请严格执行，并在最后一行明确写：'本步执行成功' 或 '本步执行失败'。"
                    )

                    original_max = self.agent.agent_config.max_steps
                    self.agent.agent_config.max_steps = 2

                    try:
                        self.agent.run(per_step_prompt)
                    except Exception as e:
                        print(f"❌ 步骤 {step_idx + 1} 执行异常: {e}")
                    finally:
                        self.agent.agent_config.max_steps = original_max

                    time.sleep(2.0)

                    step_output = redirector.get_and_clear_buffer()

                    if self._is_step_success_by_text(step_output):
                        self.current_case_step_count += 1
                        ts = datetime.now().strftime('%H:%M:%S')
                        msg = f"[{ts}] {cid} -> 步骤 {self.current_case_step_count} 通过 ✅\n"
                        self.root.after(0, lambda txt=msg: [
                            self.case_area.insert(tk.END, txt),
                            self.case_area.see(tk.END)
                        ])
                    else:
                        ts = datetime.now().strftime('%H:%M:%S')
                        msg = f"[{ts}] {cid} -> 步骤 {step_idx + 1} 失败（AI报告未找到或无法执行） ❌\n"
                        self.root.after(0, lambda txt=msg: [
                            self.case_area.insert(tk.END, txt),
                            self.case_area.see(tk.END)
                        ])

                        for r in range(step_idx + 1, len(steps_list)):
                            ts = datetime.now().strftime('%H:%M:%S')
                            umsg = f"[{ts}] {cid} -> 步骤 {r + 1} 未执行（前步失败） ❌\n"
                            self.root.after(0, lambda txt=umsg: [
                                self.case_area.insert(tk.END, txt),
                                self.case_area.see(tk.END)
                            ])
                        break

                ts = datetime.now().strftime('%H:%M:%S')
                if self.current_case_step_count == len(steps_list):
                    summary = f"[{ts}] {cid} -> 用例执行成功 🎉\n"
                else:
                    summary = f"[{ts}] {cid} -> 用例失败（部分步骤未通过） ❌\n"
                self.root.after(0, lambda txt=summary: [
                    self.case_area.insert(tk.END, txt),
                    self.case_area.see(tk.END)
                ])

                if hasattr(self.agent, 'reset') and self.agent is not None:
                    self.agent.reset()

                progress_percent = ((idx + 1) / total) * 100
                self.root.after(0, lambda p=progress_percent: self.progress_val.set(p))

        except Exception as e:
            print(f"❌ 运行时严重异常: {e}")
            import traceback
            traceback.print_exc()
        finally:
            sys.stdout = original_stdout
            if self.stop_timer:
                self.stop_timer.cancel()
                self.stop_timer = None
            if self.agent is not None:
                self.root.after(0, self._on_finish)
            elif self.btn_go['state'] == 'disabled':
                self.root.after(0, self._reset_ui_to_stopped)

    def _on_finish(self):
        if self.iproxy_process:
            self.iproxy_process.terminate()
            self.iproxy_process = None
        self.btn_go.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.btn_export.config(state="normal")
        if self.is_manual_stop:
            self.status_text_lbl.config(text="已停止", fg="orange")
        else:
            self.status_text_lbl.config(text="已完成", fg="green")
            self.progress_val.set(100)
        print("✅ 所有任务结束")


if __name__ == "__main__":
    root = tk.Tk()
    app = WindowsClassicAutoGLM(root)
    root.mainloop()